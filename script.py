import os
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import pandas as pd
folder_path = "excel_files"
def start(output_path):
    if not output_path.endswith(".xlsx"):
        output_path = output_path + ".xlsx"
    files = []
    all_data = []
    for file in os.listdir(folder_path):
        if file.endswith(".xlsx"):
            files.append(file)

    try:
        for file in files:
            file_path = os.path.join(folder_path, file)

            df = pd.read_excel(file_path, engine= "openpyxl")
            df["Source File"] = file #For adding new column
            all_data.append(df)
    except Exception as c:
        print(f"Error! Please enter files in 'excel_files' folder. {c}")
        return

    merged_df = pd.concat(all_data, ignore_index=True)
    summary_df = merged_df.groupby(["Name", "Month"]).agg({
        "Sales" : "sum"
    }).reset_index()

    merged_df.to_excel(output_path, index=False)

    workbook = openpyxl.load_workbook(output_path)

    summary_sheet = workbook.create_sheet("Summary Sheet")
    sheet = workbook.active
    sheet.title = "Merged_Data"
    thin = Side(style="thin")
    border_thin = Border(top=thin, bottom=thin, left=thin, right=thin)
    thick = Side(style="thick")
    border_thick = Border(top=thick, bottom=thick, left=thick, right=thick)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.border = border_thick
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            cell.border = border_thin
    for col in sheet.iter_cols(min_col=2, max_col=2, min_row=2):
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_sheet.append(["Name", "Month", "Sales"])
    for row in summary_df.itertuples(index=False):
        summary_sheet.append(list(row))

    sheet.column_dimensions['D'].width = 14
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.border = border_thick
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for row in summary_sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            cell.border = border_thin
    for col in summary_sheet.iter_cols(min_col=2, max_col=2, min_row=2):
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_sheet.column_dimensions['D'].width = 14
    workbook.save(output_path)
    workbook.close()
    print("Done")

if __name__ == "__main__":
    a = str(input("Enter Output Path: "))
    start(a)